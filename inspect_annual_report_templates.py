from pathlib import Path
from openpyxl import load_workbook
import json

files = [
    Path("/Users/setsuiken/Desktop/工作/学生管理系统/年度終了報告-模板/リスト.xlsx"),
    Path("/Users/setsuiken/Desktop/工作/学生管理系统/年度終了報告-模板/報告様式.xlsx"),
]

summary = []

for path in files:
    wb = load_workbook(path, data_only=False)
    book = {"file": str(path), "sheets": []}
    for ws in wb.worksheets:
        values = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                value = str(cell.value).strip()
                if value:
                    values.append({"cell": cell.coordinate, "value": value})
        book["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
            "values": values[:300],
        })
    summary.append(book)

Path("annual-report-template-inspection.json").write_text(
    json.dumps(summary, ensure_ascii=False, indent=2),
    encoding="utf-8",
)

for book in summary:
    print("\n==", Path(book["file"]).name)
    for sheet in book["sheets"]:
        print(f"-- {sheet['title']} ({sheet['max_row']}x{sheet['max_column']})")
        for item in sheet["values"][:140]:
            print(f"{item['cell']}: {item['value']}")
