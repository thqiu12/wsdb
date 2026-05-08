from pathlib import Path
from openpyxl import load_workbook
import json

path = Path("/Users/setsuiken/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/thqiu12_402a/temp/drag/WSDB Data export.xlsx")
wb = load_workbook(path, read_only=True, data_only=True)

summary = {
    "file": str(path),
    "sheets": []
}

for ws in wb.worksheets:
    rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(row)
        if idx >= 12:
            break
    non_empty_rows = []
    for idx, row in enumerate(rows, start=1):
        values = ["" if v is None else str(v) for v in row]
        if any(v.strip() for v in values):
            non_empty_rows.append({"row": idx, "values": values[:80]})

    header_candidates = []
    for row_info in non_empty_rows[:8]:
        filled = [v for v in row_info["values"] if v.strip()]
        if len(filled) >= 3:
            header_candidates.append(row_info)

    summary["sheets"].append({
        "title": ws.title,
        "max_row": ws.max_row or "unknown",
        "max_column": ws.max_column or "unknown",
        "sample_rows": non_empty_rows,
        "header_candidates": header_candidates[:3],
    })

out = Path("wsdb-export-inspection.json")
out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

print(f"File: {path.name}")
for sheet in summary["sheets"]:
    print(f"\n== {sheet['title']} ({sheet['max_row']} rows x {sheet['max_column']} cols)")
    for row in sheet["sample_rows"][:6]:
        compact = [v for v in row["values"] if v.strip()]
        print(f"R{row['row']}: " + " | ".join(compact[:30]))
