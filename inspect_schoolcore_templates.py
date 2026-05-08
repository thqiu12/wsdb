from pathlib import Path
from openpyxl import load_workbook
import json
import re

FILES = [
    Path("/Users/setsuiken/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/thqiu12_402a/temp/drag/SGG面试申请表(8).xlsx"),
    Path("/Users/setsuiken/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/thqiu12_402a/temp/drag/SGG高田马场愿书 - 学生姓名.xlsx"),
]

def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = re.sub(r"\s+", " ", v.strip())
        return v or None
    return str(v)

def inspect(path):
    wb = load_workbook(path, data_only=False, read_only=False)
    result = {"file": str(path), "sheets": []}
    for ws in wb.worksheets:
        cells = []
        labels = []
        for row in ws.iter_rows():
            for cell in row:
                value = norm(cell.value)
                if value is None:
                    continue
                cells.append({"cell": cell.coordinate, "value": value})
                if isinstance(value, str) and len(value) <= 80:
                    labels.append({"cell": cell.coordinate, "value": value})
        result["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "first_values": cells[:160],
            "labels": labels[:220],
        })
    return result

summary = [inspect(path) for path in FILES]
out = Path("template-inspection-summary.json")
out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

for book in summary:
    print("\n==", Path(book["file"]).name)
    for sheet in book["sheets"]:
        print(f"-- {sheet['title']} ({sheet['max_row']}x{sheet['max_column']})")
        for item in sheet["labels"][:80]:
            print(f"{item['cell']}: {item['value']}")
