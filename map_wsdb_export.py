from pathlib import Path
from openpyxl import load_workbook
import json

path = Path("/Users/setsuiken/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/thqiu12_402a/temp/drag/WSDB Data export.xlsx")
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb.worksheets[0]

rows = ws.iter_rows(values_only=True)
headers = next(rows)
headers = ["" if h is None else str(h).strip() for h in headers]

sample_rows = []
for idx, row in enumerate(rows, start=2):
    if idx > 7:
        break
    sample_rows.append(["" if v is None else str(v) for v in row])

print("Header count:", len(headers))
for i, h in enumerate(headers, start=1):
    if h:
        print(f"{i:03d}: {h}")

out = {
    "headers": [{"index": i, "name": h} for i, h in enumerate(headers, start=1) if h],
    "sample_rows": sample_rows,
}
Path("wsdb-export-fields.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
